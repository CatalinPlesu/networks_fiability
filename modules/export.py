import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
import os
import random

yellow = "ffff6d"
red = "ff6d6d"
green = "afd095"

file_prefix = random.random()
output_dir = "output"

try:
    os.mkdir(output_dir)
    print(f" folderul '{output_dir}' a fost creata")
except:
    print(f"nu a putut fi creat folderul'{output_dir}'")
    pass


def file_path(distribution: str):
    return os.path.join(output_dir, f"{distribution}_{file_prefix}.xlsx")

def create_workbook(distribution: str):
    wb = openpyxl.Workbook() 
    ws = wb.active
    ws = wb['Sheet']
    ws.title = "diff"
    ws_sp = wb.create_sheet("sp") 
    ws_ps = wb.create_sheet("ps") 
    ws_fav = wb.create_sheet("fav") 
    ws.sheet_properties.tabColor = "dba281"
    ws_sp.sheet_properties.tabColor = "99c39e"
    ws_ps.sheet_properties.tabColor = "89cdd3"
    ws_fav.sheet_properties.tabColor = "d3869b"

    global file_prefix 
    file_prefix = random.random()

    wb.save(filename = file_path(distribution))
    return wb

# wb = create_workbook()

def include_data(sheet, workbook, distribution, matrix):
    for row in matrix:
        workbook[sheet].append(row)
    workbook.save(filename = file_path(distribution))

def pretty_output(sheet, workbook, distribution, sp, ps, fav):
    include_data('sp', workbook, distribution, sp)
    include_data('ps', workbook, distribution, ps)
    include_data('fav', workbook, distribution, fav)
    diff_sheet(sheet, workbook, distribution)

def diff_sheet(sheet, workbook, distribution):
    for row in workbook[sheet]:
        for cell in row:
            workbook['diff'][cell.coordinate].value = cell.value
    
    for (col, col_ps, col_sp, col_fav) in zip(workbook['diff'], workbook['ps'], workbook['sp'], workbook['fav']):
        for (cell, cell_ps, cell_sp, cell_fav) in zip(col, col_ps, col_sp, col_fav):
           try:
               ps = int(float(cell_ps.value))
               sp = int(float(cell_sp.value))
           except:
               ps = 0
               sp = 0
           if ps > 255:
               ps = 255
           if sp > 255:
               sp = 255
           cell_ps.fill = PatternFill(start_color = '%02x%02x%02x' % (255 - ps, 255 - ps, 255), fill_type = "solid")
           cell_sp.fill = PatternFill(start_color = '%02x%02x%02x' % (255 - sp, 255, 255 - sp), fill_type = "solid")
           if cell_ps.value == cell_sp.value:#colorize cell yellow
               cell.fill = PatternFill(start_color = yellow, fill_type = "solid")
               # cell_fav.fill = PatternFill(start_color = yellow, fill_type = "solid")
           elif cell_ps.value > cell_sp.value:#colorize cell red
               cell.fill = PatternFill(start_color = red, fill_type = "solid")
               if cell_fav.value == "P":
                   cell_fav.fill = PatternFill(start_color = yellow, fill_type = "solid")
           elif cell_ps.value < cell_sp.value:#colorize cell green
               cell.fill = PatternFill(start_color = green, fill_type = "solid")
               if cell_fav.value == "S":
                   cell_fav.fill = PatternFill(start_color = red, fill_type = "solid")
           if cell_fav.value == "B":
               cell_fav.fill = PatternFill(start_color = green, fill_type = "solid")
    workbook.save(filename = file_path(distribution))

# diff_sheet('ps', wb)
