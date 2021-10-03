import openpyxl
# import process_data
from openpyxl.styles import PatternFill
from datetime import datetime
import csv

file_prefix = datetime.now().strftime("%d-%m-%Y_%H:%M:%S_")
output_dir = "output"

def file_path(file = ""):
    return output_dir + "/"+file_prefix + file + ".xlsx"

def create_workbook(version = 'a'):
    wb = openpyxl.Workbook() 
    ws = wb.active
    ws = wb['Sheet']
    ws.title = "diff"
    ws_sp = wb.create_sheet("sp") 
    ws_ps = wb.create_sheet("ps") 
    ws.sheet_properties.tabColor = "dba281"
    ws_sp.sheet_properties.tabColor = "99c39e"
    ws_ps.sheet_properties.tabColor = "89cdd3"
    wb.save(filename = file_path(version))
    return wb

# wb = create_workbook()

def load_workbook(version = 'a'):
    return openpyxl.load_workbook(file_path('a'))

def csv_to_sheet(sheet, workbook, version = 'a'):
    with open(output_dir + '/' + sheet + '.csv') as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            workbook[sheet].append(row)
    workbook.save(filename = file_path(version))

def pretty_output(sheet, workbook, version = 'a'):
    csv_to_sheet('sp', workbook, version)
    csv_to_sheet('ps', workbook, version)
    diff_sheet(sheet, workbook, version)

def export_row(row, sheet, workbook, version = 'a'):
    workbook[sheet].append(row)
    workbook.save(filename = file_path(version))

# export_row([1,2,3,4,5], 'sp', wb)

def diff_sheet(sheet, workbook, version = 'a'):
    for row in workbook[sheet]:
        for cell in row:
            workbook['diff'][cell.coordinate].value = cell.value
    
    for (col, col_ps, col_sp) in zip(workbook['diff'], workbook['ps'], workbook['sp']):
        for (cell, cell_ps, cell_sp) in zip(col, col_ps, col_sp):
           ps = int(cell_ps.value) 
           sp = int(cell_sp.value)
           cell_ps.fill = PatternFill(start_color='%02x%02x%02x' % (255 - ps, 255 - ps, 255), fill_type = "solid")
           cell_sp.fill = PatternFill(start_color='%02x%02x%02x' % (255 - sp, 255, 255 - sp), fill_type = "solid")
           if cell_ps.value == cell_sp.value:#colorize cell yellow
               cell.fill = PatternFill(start_color="ffff6d", fill_type = "solid")
           if cell_ps.value > cell_sp.value:#colorize cell red
               cell.fill = PatternFill(start_color="ff6d6d", fill_type = "solid")
           if cell_ps.value < cell_sp.value:#colorize cell green
               cell.fill = PatternFill(start_color="afd095", fill_type = "solid")
    workbook.save(filename = file_path(version))



# diff_sheet('ps', wb)
