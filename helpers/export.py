import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
import os
import random

from helpers.json_config import load_settings
sett = load_settings()

file_identifier = random.random()

try:
    os.mkdir(sett['output']['dir'])
    print(f" folderul '{sett['output']['dir']}' a fost creata")
except:
    # print(f"nu a putut fi creat folderul'{sett['output']['dir']}'")
    pass

def file_path(distribution: str):
    return os.path.join(sett['output']['dir'], f"{distribution}_{file_identifier}.xlsx")

def create_workbook(distribution: str):
    wb = openpyxl.Workbook() 
    ws = wb.active
    ws = wb['Sheet']
    ws.title = "theorem"
    ws_sp = wb.create_sheet("sp") 
    ws_ps = wb.create_sheet("ps") 
    ws.sheet_properties.tabColor = "dba281"
    ws_sp.sheet_properties.tabColor = "99c39e"
    ws_ps.sheet_properties.tabColor = "89cdd3"

    global file_identifier 
    file_identifier = random.random()

    wb.save(filename = file_path(distribution))
    return wb

# wb = create_workbook()

def include_data(sheet, workbook, distribution, matrix):
    for row in matrix:
        workbook[sheet].append(row)
    workbook.save(filename = file_path(distribution))

def pretty_output(workbook, distribution, sp, ps, theorem):
    include_data('sp', workbook, distribution, sp)
    include_data('ps', workbook, distribution, ps)
    include_data('theorem', workbook, distribution, theorem)
    diff_sheet(workbook, distribution)

def diff_sheet(workbook, distribution):
    for row in workbook['theorem']:
        for cell in row:
            workbook['theorem'][cell.coordinate].value = cell.value
    
    for (col, col_ps, col_sp) in zip(workbook['theorem'], workbook['ps'], workbook['sp']):
        for (cell, cell_ps, cell_sp) in zip(col, col_ps, col_sp):
           try:
               ps = int(float(cell_ps.value))
               sp = int(float(cell_sp.value))
           except:
               ps = 0
               sp = 0
           if ps > 255:
               ps = 255
           if sp > 255:
               sp = 255

           cell_ps.fill = PatternFill(start_color = '%02x%02x%02x' % (255 - ps, 255 - ps, 255), fill_type = "solid")
           cell_sp.fill = PatternFill(start_color = '%02x%02x%02x' % (255 - sp, 255, 255 - sp), fill_type = "solid")

           if cell_ps.value == cell_sp.value:
               cell.fill = PatternFill(start_color = sett['output']['colors']['yellow'], fill_type = "solid")

           elif cell_ps.value > cell_sp.value:
               cell.fill = PatternFill(start_color = sett['output']['colors']['red'], fill_type = "solid")
               if cell.value == "P":
                   cell.fill = PatternFill(start_color = sett['output']['colors']['yellow'], fill_type = "solid")

           elif cell_ps.value < cell_sp.value:
               cell.fill = PatternFill(start_color = sett['output']['colors']['green'], fill_type = "solid")
               if cell.value == "S":
                   cell.fill = PatternFill(start_color = sett['output']['colors']['red'], fill_type = "solid")

           if cell.value == "B":
               cell.fill = PatternFill(start_color = sett['output']['colors']['green'], fill_type = "solid")
    workbook.save(filename = file_path(distribution))

# diff_sheet('ps', wb)
