# font_sizes.py

import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import BLUE as YELLOW

def font_demo(path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    cell = sheet["A1"]
    sheet['A1'].fill = PatternFill(start_color="FFC7CE", fill_type = "solid")
    sheet['G8'].fill = PatternFill(fgColor=YELLOW, fill_type = "solid")
    cell.font = Font(size=12)
    cell.value = "Hello"

    cell2 = sheet["A2"]
    cell2.font = Font(name="Arial", size=14, color="00FF0000")
    sheet["A2"] = "from"

    cell2 = sheet["A3"]
    cell2.font = Font(name="Tahoma", size=16, color="00339966")
    sheet["A3"] = "OpenPyXL"

    workbook.save(path)


if __name__ == "__main__":
    font_demo("font_demo.xlsx")
