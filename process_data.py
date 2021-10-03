import csv
import openpyxl
from openpyxl.styles import PatternFill


def get_matrix(file):
    file = open(file + '.csv')
    data = csv.reader(file)
    return list(data)

wb = openpyxl.Workbook() 
ws = wb.active
ws = wb['Sheet']
ws.title = 'comparatie'
ws_sp = wb.create_sheet("sp") 
ws_ps = wb.create_sheet("ps") 
ws.sheet_properties.tabColor = "dba281"
ws_sp.sheet_properties.tabColor = "99c39e"
ws_ps.sheet_properties.tabColor = "89cdd3"

ps = get_matrix('output/ps')
for row in ps:
    ws_ps.append(row)

sp = get_matrix('output/sp')
for row in sp:
    ws_sp.append(row)
    ws.append(row)

for (col, col_ps, col_sp) in zip(ws.iter_cols(), ws_ps.iter_cols(), ws_sp.iter_cols()):
    for (cell, cell_ps, cell_sp) in zip(col, col_ps, col_sp):
       # cell.value = cell_ps.value
       ps = int(cell_ps.value) 
       sp = int(cell_sp.value)
       cell_ps.fill = PatternFill(start_color='%02x%02x%02x' % (255 - ps, 255 - ps, 255), fill_type = "solid")
       cell_sp.fill = PatternFill(start_color='%02x%02x%02x' % (255 - sp, 255, 255 - sp), fill_type = "solid")
       if cell_ps.value == cell_sp.value:#colorize cell yellow
           cell.fill = PatternFill(start_color="ffff6d", fill_type = "solid")
       if cell_ps.value > cell_sp.value:#colorize cell red
           cell.fill = PatternFill(start_color="ff6d6d", fill_type = "solid")
       if cell_ps.value < cell_sp.value:#colorize cell green
           cell.fill = PatternFill(start_color="afd095", fill_type = "solid")

wb.save(filename='Test.xlsx')
